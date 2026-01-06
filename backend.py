import logging
import os
import secrets
import datetime
from typing import Dict, List, Optional
from uuid import uuid4

import pandas as pd
import torch
import openpyxl
from openpyxl import Workbook
from fastapi import FastAPI, HTTPException, Request
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sentence_transformers import SentenceTransformer, util

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------------------------------------------------------------
# GLOBAL CONF & MODEL LOADING
# -------------------------------------------------------------------------
MODEL_PATH = "finetuned_chatbot_model"
EXCEL_PATH = "notebooks/WITDS_FAQ_Paraphrased.xlsx"  # Adjusted path based on typical structure, checks below
# Actually, looking at main.py: "C:/Users/BIBHAV KUMAR/Desktop/internship/New_Chatbot_with_pridiction/WITDS_FAQ_Paraphrased.xlsx"
# But list_dir showed it in the current dir.
EXCEL_PATH_LOCAL = "WITDS_FAQ_Paraphrased.xlsx"

model: Optional[SentenceTransformer] = None
faq_df: Optional[pd.DataFrame] = None
faq_embeddings: Optional[torch.Tensor] = None
faq_questions: List[str] = []
faq_answers: List[str] = []
next_question_map: Dict[str, str] = {}

IRRELEVANT_THRESHOLD = 0.5
MAX_IRRELEVANT_LIMIT = 3
TIMEOUT_SECONDS = 300  # For session expiration check if needed

class SessionState:
    def __init__(self):
        self.previous_questions: List[str] = []
        self.awaiting_confirmation: bool = False
        self.irrelevant_count: int = 0
        self.session_ended: bool = False
        self.last_activity: datetime.datetime = datetime.datetime.now()

# In-memory session store
sessions: Dict[str, SessionState] = {}

# -------------------------------------------------------------------------
# INITIALIZATION
# -------------------------------------------------------------------------
@app.on_event("startup")
def startup_event():
    global model, faq_df, faq_embeddings, faq_questions, faq_answers, next_question_map
    
    print("Loading model...")
    # Use local path if exists, else trigger error or download? 
    # main.py used "finetuned_chatbot_model"
    if os.path.exists(MODEL_PATH):
        model = SentenceTransformer(MODEL_PATH)
    else:
        # Fallback or try to load sentence-transformers default? 
        # Assuming the user has it since they ran main.py
        print(f"Warning: {MODEL_PATH} not found, trying default 'all-MiniLM-L6-v2' just in case or failing.")
        model = SentenceTransformer("all-MiniLM-L6-v2") # Fallback for now to ensure start

    print("Loading Excel data...")
    if os.path.exists(EXCEL_PATH_LOCAL):
        path = EXCEL_PATH_LOCAL
    else:
        # Fallback to absolute path from main.py if local not found
        path = "C:/Users/BIBHAV KUMAR/Desktop/internship/New_Chatbot_with_pridiction/WITDS_FAQ_Paraphrased.xlsx"
    
    try:
        faq_df = pd.read_excel(path)
        faq_df = faq_df.dropna(subset=["Question", "Answer"]).fillna("")
        faq_df["Question"] = faq_df["Question"].astype(str)
        faq_df["Answer"] = faq_df["Answer"].astype(str)
        faq_df["Next Question"] = faq_df["Next Question"].astype(str)

        faq_questions = faq_df["Question"].tolist()
        faq_answers = faq_df["Answer"].tolist()
        faq_embeddings = model.encode(faq_questions, convert_to_tensor=True)
        next_question_map = dict(zip(faq_df["Question"], faq_df["Next Question"]))
        print("Model and Data Loaded Successfully.")
    except Exception as e:
        print(f"Error loading Excel: {e}")

# -------------------------------------------------------------------------
# HELPER FUNCTIONS
# -------------------------------------------------------------------------
def ensure_excel_file(path, headers):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(path)

def save_contact_to_excel(name, email, phone, path="user_contacts.xlsx"):
    ensure_excel_file(path, ["Name", "Email", "Phone", "Timestamp"])
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.append([name, email, phone, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(path)

def save_feedback_to_excel(feedback, path="user_feedback.xlsx"):
    ensure_excel_file(path, ["Feedback", "Timestamp"])
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.append([feedback, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(path)

def get_session(session_id: str) -> SessionState:
    if session_id not in sessions:
        sessions[session_id] = SessionState()
    # Check timeout?
    sess = sessions[session_id]
    sess.last_activity = datetime.datetime.now()
    return sess

def get_answer(query):
    query_vec = model.encode(query, convert_to_tensor=True)
    sims = util.cos_sim(query_vec, faq_embeddings)
    idx = torch.argmax(sims)
    best_score = sims[0][idx].item()
    return faq_answers[idx], faq_questions[idx], best_score

def is_duplicate(query, session: SessionState, threshold=0.8):
    query_vec = model.encode(query, convert_to_tensor=True)
    for prev_q in session.previous_questions:
        prev_vec = model.encode(prev_q, convert_to_tensor=True)
        sim = util.cos_sim(query_vec, prev_vec).item()
        if sim > threshold:
            return True
    session.previous_questions.append(query)
    return False

def get_related_questions(query, top_k=3):
    query_vec = model.encode(query, convert_to_tensor=True)
    sims = util.cos_sim(query_vec, faq_embeddings)[0]
    top_results = torch.topk(sims, k=top_k + 1)
    related_qs = []
    for score, idx in zip(top_results.values, top_results.indices):
        question = faq_questions[idx]
        if question != query and question not in related_qs:
            related_qs.append(question)
        if len(related_qs) == top_k:
            break
    return related_qs

# -------------------------------------------------------------------------
# API MODELS
# -------------------------------------------------------------------------
class ChatRequest(BaseModel):
    query: str
    session_id: Optional[str] = None

class FeedbackRequest(BaseModel):
    feedback: str

class ContactRequest(BaseModel):
    name: str
    email: str
    phone: Optional[str] = ""

class ChatResponse(BaseModel):
    response: str
    session_id: str
    session_ended: bool
    suggestions: List[str] = []

# -------------------------------------------------------------------------
# ENDPOINTS
# -------------------------------------------------------------------------
@app.post("/chat", response_model=ChatResponse)
async def chat_endpoint(request: ChatRequest):
    if not request.session_id:
        request.session_id = str(uuid4())
    
    session = get_session(request.session_id)
    query = request.query.strip()

    if session.session_ended:
         return ChatResponse(
            response="The session has ended. Please refresh to start a new chat.",
            session_id=request.session_id,
            session_ended=True
        )

    # 1. Handle Confirmation Logic
    if session.awaiting_confirmation:
        if query.lower() in ["yes", "haan", "hn", "yes it is", "y"]:
            session.session_ended = True
            msg = "I'm glad I could help! 🌟\n\"Helping one person might not change the whole world, but it could change the world for one person.\"\nPlease share your details below before you go."
            return ChatResponse(response=msg, session_id=request.session_id, session_ended=True)
        elif query.lower() in ["no", "nahi", "nah", "n"]:
            session.awaiting_confirmation = False
            return ChatResponse(response="Okay, feel free to ask your next question.", session_id=request.session_id, session_ended=False)
        else:
            return ChatResponse(response="Please respond with 'yes' or 'no'.", session_id=request.session_id, session_ended=False)

    # 2. Get Answer
    answer, matched_question, best_score = get_answer(query)

    # 3. Check Relevance
    if best_score < IRRELEVANT_THRESHOLD:
        session.irrelevant_count += 1
        msg = f"🤔 I'm not sure about that. (Confidence: {best_score:.2f})\nPlease try rephrasing your question."
        if session.irrelevant_count > MAX_IRRELEVANT_LIMIT:
            session.session_ended = True
            msg = "❌ Too many irrelevant queries detected. Session ended."
        return ChatResponse(response=msg, session_id=request.session_id, session_ended=session.session_ended)
    
    session.irrelevant_count = 0 

    # 4. Check Duplicate
    if is_duplicate(query, session):
        session.awaiting_confirmation = True
        return ChatResponse(
            response="You've already asked this or a similar question.\nDid that answer resolve your query? (yes/no)",
            session_id=request.session_id,
            session_ended=False
        )

    # 5. Determine Next Steps (Related/Next Q)
    next_q = next_question_map.get(matched_question, "").strip() or None
    suggestions = []
    
    final_response = answer
    if next_q:
        final_response += f"\n\nWould you like to know: {next_q}"
        suggestions.append(next_q)
    else:
        related = get_related_questions(query)
        if related:
            suggestions = related
    
    return ChatResponse(
        response=final_response,
        session_id=request.session_id,
        session_ended=False,
        suggestions=suggestions
    )

@app.post("/feedback")
async def feedback_endpoint(request: FeedbackRequest):
    save_feedback_to_excel(request.feedback)
    return {"status": "success", "message": "Feedback saved"}

@app.post("/contact")
async def contact_endpoint(request: ContactRequest):
    save_contact_to_excel(request.name, request.email, request.phone)
    return {"status": "success", "message": "Contact details saved"}

# Mount static files
app.mount("/", StaticFiles(directory="static", html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    # If running directly
    uvicorn.run(app, host="127.0.0.1", port=8000)
