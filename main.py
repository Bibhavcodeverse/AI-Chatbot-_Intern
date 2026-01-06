from sentence_transformers import SentenceTransformer, util
import pandas as pd
import torch
import os
import datetime
import openpyxl
from openpyxl import Workbook
from inputimeout import inputimeout, TimeoutOccurred

# Load model and data
model = SentenceTransformer("finetuned_chatbot_model")
faq_df = pd.read_excel("C:/Users/BIBHAV KUMAR/Desktop/internship/New_Chatbot_with_pridiction/WITDS_FAQ_Paraphrased.xlsx")
faq_df = faq_df.dropna(subset=["Question", "Answer"]).fillna("")
faq_df["Question"] = faq_df["Question"].astype(str)
faq_df["Answer"] = faq_df["Answer"].astype(str)
faq_df["Next Question"] = faq_df["Next Question"].astype(str)

faq_questions = faq_df["Question"].tolist()
faq_answers = faq_df["Answer"].tolist()
faq_embeddings = model.encode(faq_questions, convert_to_tensor=True)
next_question_map = dict(zip(faq_df["Question"], faq_df["Next Question"]))

# Session state
previous_questions = []
awaiting_confirmation = False
greeting_sent = False
session_ended = False
TIMEOUT_SECONDS = 300
irrelevant_count = 0
IRRELEVANT_THRESHOLD = 0.5
MAX_IRRELEVANT_LIMIT = 3

# Excel saving utilities
def ensure_excel_file(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Email", "Phone", "Timestamp"])
        wb.save(path)

def save_contact_to_excel(name, email, phone, path="user_contacts.xlsx"):
    ensure_excel_file(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.append([name, email, phone, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(path)

def ensure_feedback_file(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Feedback", "Timestamp"])
        wb.save(path)

def save_feedback_to_excel(feedback, path="user_feedback.xlsx"):
    ensure_feedback_file(path)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.append([feedback, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(path)

def collect_feedback():
    print("üìù Before you go, we'd love your feedback!")
    try:
        feedback = inputimeout(prompt="üí¨ Your Feedback: ", timeout=120)
        save_feedback_to_excel(feedback)
        print("üôè Thank you for your feedback!")
    except TimeoutOccurred:
        print("‚è≥ No feedback received (timeout). Exiting.")

def collect_contact_details():
    global session_ended
    print("üìû Please share your contact details before exiting.")
    try:
        name = inputimeout(prompt="üë§ Your Name: ", timeout=60)
        email = inputimeout(prompt="üìß Your Email: ", timeout=60)
        phone = inputimeout(prompt="üìû Your Phone (optional): ", timeout=60)
        save_contact_to_excel(name, email, phone)
        print("‚úÖ Bot: Thank you! Your details have been saved.")
        session_ended = True
    except TimeoutOccurred:
        print("‚è≥ Contact input timed out.")

# Chatbot logic
def is_duplicate(query, threshold=0.8):
    query_vec = model.encode(query, convert_to_tensor=True)
    for prev_q in previous_questions:
        prev_vec = model.encode(prev_q, convert_to_tensor=True)
        sim = util.cos_sim(query_vec, prev_vec).item()
        if sim > threshold:
            return True
    previous_questions.append(query)
    return False

def get_answer(query):
    query_vec = model.encode(query, convert_to_tensor=True)
    sims = util.cos_sim(query_vec, faq_embeddings)
    idx = torch.argmax(sims)
    best_score = sims[0][idx].item()
    return faq_answers[idx], faq_questions[idx], best_score

def get_next_question(matched_question):
    return next_question_map.get(matched_question, "").strip() or None

def get_related_questions(query, top_k=3):
    query_vec = model.encode(query, convert_to_tensor=True)
    sims = util.cos_sim(query_vec, faq_embeddings)[0]
    top_results = torch.topk(sims, k=top_k + 1)
    related_qs = []
    for score, idx in zip(top_results.values, top_results.indices):
        question = faq_questions[idx]
        if question != query and question not in related_qs:
            related_qs.append(question)
        if len(related_qs) == top_k:
            break
    return related_qs

def chatbot(query):
    global awaiting_confirmation, greeting_sent, session_ended, irrelevant_count

    if session_ended:
        return "The session has ended. Thank you!"

    # ‚úÖ Handle duplicate confirmation FIRST ‚Äî before embedding logic
    if awaiting_confirmation:
        if query.lower() in ["yes", "haan", "hn", "yes it is"]:
            print("Bot: I'm glad I could help! üåü")
            print('Bot: "Helping one person might not change the whole world, but it could change the world for one person."')
            collect_contact_details()
            collect_feedback()
            print("üîí Bot: Session ended after successful interaction.")
            session_ended = True
            raise SystemExit
        elif query.lower() in ["no", "nahi", "nah"]:
            awaiting_confirmation = False
            return "Okay, feel free to ask your next question."
        else:
            return "Please respond with 'yes' or 'no'."

    # üîΩ All logic below only runs if not in confirmation mode

    # Get answer + score
    answer, matched_question, best_score = get_answer(query)

    # Handle irrelevant question
    if best_score < IRRELEVANT_THRESHOLD:
        irrelevant_count += 1
        print(f"ü§î This seems unrelated to our FAQ topics. (Score: {best_score:.2f})")
        print("Bot: Please try rephrasing your question.")
        if irrelevant_count > MAX_IRRELEVANT_LIMIT:
            print("‚ùå Bot: Too many irrelevant queries detected.")
            collect_feedback()
            collect_contact_details()
            print("üîí Bot: Session ended after repeated irrelevant queries.")
            raise SystemExit
        return ""

    irrelevant_count = 0

    # Handle duplicate question
    if is_duplicate(query):
        awaiting_confirmation = True
        return "You've already asked this or a similar question.\nDid that answer resolve your query? (yes/no)"

    # Continue normal flow
    next_q = get_next_question(matched_question)
    if next_q:
        return f"{answer}\nWould you like to know: {next_q}"
    else:
        related_qs = get_related_questions(query)
        if related_qs:
            suggestions = "\n".join([f"- {q}" for q in related_qs])
            return f"{answer}\nYou may also be interested in:\n{suggestions}"
        return answer

# CLI chatbot loop
if __name__ == "__main__":
    print("ü§ñ Chatbot loaded. Ask a question or type 'exit' to quit.")
    while True:
        if session_ended:
            print("üîí Bot: Session ended. No further input will be accepted.")
            raise SystemExit

        try:
            user_input = inputimeout(prompt="You: ", timeout=TIMEOUT_SECONDS)
        except TimeoutOccurred:
            print("‚è≥ Bot: Session ended due to inactivity (timeout).")
            collect_feedback()
            raise SystemExit

        if user_input.lower() in ["exit", "quit"]:
            print("Bot: Goodbye!")
            collect_feedback()
            collect_contact_details()
            raise SystemExit

        response = chatbot(user_input)
        if response:
            print("Bot:", response)

        if greeting_sent:
            collect_contact_details()
